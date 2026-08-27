"""
CONCILPRO — API de Conciliação de Razão de Fornecedores.

Endpoints montados em /api/v1/empresas/{empresa_id}/concilpro/*.

Endpoints de query usam AsyncSession do FastAPI.
O processamento pesado (parsing + conciliação) roda em BackgroundTasks com
SyncSessionLocal (psycopg2) para compatibilidade com o algoritmo FIFO síncrono.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from decimal import Decimal
from typing import Literal, Optional
from uuid import UUID

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.api.deps import get_company_context, require_csrf
from src.api.uploads import ler_upload_limitado
from src.api.autorizacao import requer
from src.core.errors import AppError
from src.db.models import (
    CpArquivo as ArquivoImportado,
    CpFornecedor as Fornecedor,
    CpLancamento as LancamentoFornecedor,
    CpDivergencia as Divergencia,
    CpConciliacao as ConciliacaoInterna,
)
from src.db.session import SyncSessionLocal, get_db
from src.domain.exportacao.formatos import (
    COLUNAS_LANCAMENTOS_IMPORTACAO,
    _dicts_to_csv,
    _dicts_to_txt,
    _dicts_to_xlsx,
)

logger = logging.getLogger(__name__)

# Escopo no router, não rota a rota: toda rota nova exige acesso à empresa do path.
router = APIRouter(
    prefix="/empresas/{empresa_id}/concilpro",
    tags=["concilpro"],
    dependencies=[Depends(get_company_context)],
)

# ============================================================================
# UTILITÁRIOS
# ============================================================================

def _converter_data_br(valor) -> Optional[datetime]:
    """Converte string DD/MM/YYYY → datetime; passa datetime sem alterar."""
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor
    try:
        return datetime.strptime(str(valor), "%d/%m/%Y")
    except (ValueError, TypeError):
        return None


CENTAVO = Decimal("0.01")


def _moeda(valor) -> Decimal:
    """Normaliza um valor para a precisão persistida pelo schema."""
    return Decimal(str(valor or "0")).quantize(CENTAVO)


def _saldo_anterior_assinado(valor, tipo: str | None) -> Decimal:
    """Fornecedor tem natureza credora: C aumenta a obrigação e D a reduz."""
    modulo = abs(_moeda(valor))
    return -modulo if (tipo or "").strip().upper() == "D" else modulo


def _calcular_posicao_fornecedor(
    saldo_anterior,
    saldo_anterior_tipo: str | None,
    total_debito,
    total_credito,
    tem_lancamentos: bool,
) -> tuple[Decimal, str, str]:
    """Retorna saldo assinado, tipo final e status pela mesma equação contábil."""
    abertura = _saldo_anterior_assinado(saldo_anterior, saldo_anterior_tipo)
    debitos = _moeda(total_debito)
    creditos = _moeda(total_credito)
    saldo = (abertura + creditos - debitos).quantize(CENTAVO)

    if not tem_lancamentos and abertura == 0 and debitos == 0 and creditos == 0:
        status = "SEM_MOVIMENTO"
    elif saldo == 0:
        status = "QUITADO"
    elif saldo < 0:
        status = "ADIANTADO"
    else:
        status = "EM_ABERTO"

    saldo_tipo = "C" if saldo > 0 else ("D" if saldo < 0 else "")
    return saldo, saldo_tipo, status


def _celula_texto_segura(valor: str | None) -> str | None:
    """Neutraliza texto que o Excel interpretaria como fórmula."""
    if valor is None:
        return None
    texto = str(valor)
    significativo = texto.lstrip()
    if significativo.startswith(("=", "+", "-", "@")):
        return "'" + texto
    return texto


# ============================================================================
# BACKGROUND TASK (síncrono — usa psycopg2 via SyncSessionLocal)
# ============================================================================

def _processar_arquivo_background(arquivo_id: int, empresa_id: UUID, conteudo: bytes) -> None:
    """
    Processamento pesado em background — usa sessão própria (não a da request).
    Chamado pelo BackgroundTasks do FastAPI após o upload retornar ao cliente.
    """
    from src.domain.concilpro.parser import parsear_arquivo_razao
    from src.domain.concilpro.consolidador import consolidar_todos_fornecedores
    from src.domain.concilpro.ai_classifier import classificar_lancamentos_incertos
    from src.domain.concilpro.conciliacao_intel import conciliar_todos_fornecedores_inteligente

    db = None
    try:
        db = SyncSessionLocal()
        arquivo = db.query(ArquivoImportado).filter(
            ArquivoImportado.id == arquivo_id,
            ArquivoImportado.empresa_id == empresa_id,
        ).first()
        if not arquivo:
            logger.error("❌ Background: arquivo_id=%d não encontrado", arquivo_id)
            return

        dados = parsear_arquivo_razao(conteudo)
        dados = consolidar_todos_fornecedores(dados)

        incertos = [
            lanc
            for forn in dados["fornecedores"]
            for lanc in forn.get("lancamentos", [])
            if lanc.get("classificacao_incerta")
        ]
        if incertos:
            logger.info("🤖 Enviando %d lançamentos incertos para classificação IA…", len(incertos))
            classificar_lancamentos_incertos(incertos)

        arquivo.data_inicio        = _converter_data_br(dados.get("periodo_inicio"))
        arquivo.data_fim           = _converter_data_br(dados.get("periodo_fim"))
        arquivo.empresa            = dados.get("empresa")
        arquivo.cnpj_empresa       = dados.get("cnpj")
        arquivo.total_fornecedores = dados.get("total_fornecedores", len(dados["fornecedores"]))
        arquivo.total_lancamentos  = dados.get(
            "total_lancamentos",
            sum(len(f.get("lancamentos", [])) for f in dados["fornecedores"]),
        )

        logger.info("💾 Inserindo %d fornecedores…", len(dados["fornecedores"]))

        for idx, forn_data in enumerate(dados["fornecedores"], 1):
            if idx % 50 == 0:
                logger.info("   Processados %d / %d", idx, len(dados["fornecedores"]))

            saldo_anterior = Decimal(str(forn_data.get("saldo_anterior", 0)))

            lancamentos_raw = forn_data.get("lancamentos", [])
            total_debito_calc  = sum(Decimal(str(l.get("valor_debito",  0))) for l in lancamentos_raw)
            total_credito_calc = sum(Decimal(str(l.get("valor_credito", 0))) for l in lancamentos_raw)

            total_debito_ia  = Decimal(str(forn_data.get("total_debito",  0)))
            total_credito_ia = Decimal(str(forn_data.get("total_credito", 0)))

            total_debito  = total_debito_calc  if total_debito_calc  > 0 else total_debito_ia
            total_credito = total_credito_calc if total_credito_calc > 0 else total_credito_ia

            # ── Guarda de divergência ───────────────────────────────────────
            # A soma dos lançamentos parseados vence o total declarado no arquivo
            # (acima). Quando os dois discordam, os números não são confiáveis —
            # marcar para revisão em vez de entregar silenciosamente. A mensagem
            # informa o quanto falta, que é o que orienta a conferência manual.
            TOL_DIVERGENCIA = Decimal("0.01")
            divergencias: list[str] = []

            if total_debito_ia > 0 and abs(total_debito_calc - total_debito_ia) > TOL_DIVERGENCIA:
                divergencias.append(
                    f"débito: PDF declara {total_debito_ia}, lançamentos somam {total_debito_calc}"
                )
            if total_credito_ia > 0 and abs(total_credito_calc - total_credito_ia) > TOL_DIVERGENCIA:
                divergencias.append(
                    f"crédito: PDF declara {total_credito_ia}, lançamentos somam {total_credito_calc}"
                )

            if divergencias:
                logger.warning(
                    "⚠️ Divergência em '%s' (conta %s): %s",
                    forn_data["nome_fornecedor"][:40],
                    forn_data["codigo_conta"],
                    "; ".join(divergencias),
                )

            saldo_ant_tipo = (forn_data.get("saldo_anterior_tipo") or "")[:1]
            saldo_final, saldo_final_tipo, status_pagamento = _calcular_posicao_fornecedor(
                saldo_anterior,
                saldo_ant_tipo,
                total_debito,
                total_credito,
                bool(lancamentos_raw),
            )

            fornecedor = Fornecedor(
                empresa_id          = empresa_id,
                arquivo_origem_id   = arquivo.id,
                codigo_conta        = forn_data["codigo_conta"][:10],
                conta_contabil      = forn_data["conta_contabil"][:50],
                nome_fornecedor     = forn_data["nome_fornecedor"],
                saldo_anterior      = saldo_anterior,
                saldo_anterior_tipo = saldo_ant_tipo,
                total_debito        = total_debito,
                total_credito       = total_credito,
                saldo_final         = saldo_final,
                saldo_final_tipo    = saldo_final_tipo,
                valor_a_pagar       = saldo_final,
                status_pagamento    = status_pagamento,
                divergencia_calculo = bool(divergencias),
                mensagem_erro       = "; ".join(divergencias)[:2000] or None,
            )
            db.add(fornecedor)
            db.flush()

            for lanc_data in forn_data.get("lancamentos", []):
                vd    = Decimal(str(lanc_data["valor_debito"]))
                vc    = Decimal(str(lanc_data["valor_credito"]))
                saldo = Decimal(str(lanc_data["saldo_apos_lancamento"]))

                lote_val          = (lanc_data.get("lote") or "")[:50] or None
                conta_partida_val = (str(lanc_data.get("conta_partida") or "")[:20]) or None
                saldo_tipo_val    = (lanc_data.get("saldo_tipo") or "")[:1]
                numero_nf_val     = (lanc_data.get("numero_nf") or "")[:50] or None
                tipo_op_val       = (lanc_data.get("tipo_operacao") or "OUTRO")[:20]

                db.add(LancamentoFornecedor(
                    empresa_id            = empresa_id,
                    fornecedor_id         = fornecedor.id,
                    data_lancamento       = lanc_data["data_lancamento"],
                    lote                  = lote_val,
                    historico             = lanc_data["historico"],
                    conta_partida         = conta_partida_val,
                    valor_debito          = vd,
                    valor_credito         = vc,
                    saldo_apos_lancamento = saldo,
                    saldo_tipo            = saldo_tipo_val,
                    tipo_operacao         = tipo_op_val,
                    numero_nf             = numero_nf_val,
                    cnpj_historico        = lanc_data.get("cnpj_historico"),
                    valor_saldo           = vc if lanc_data["tipo_operacao"] == "COMPRA" else Decimal("0"),
                    classificado_por_ia   = lanc_data.get("classificado_por_ia", False),
                ))

        logger.info("🔄 Iniciando conciliação inteligente…")
        conciliar_todos_fornecedores_inteligente(db, arquivo.id, empresa_id)

        for forn in db.query(Fornecedor).filter(
            Fornecedor.arquivo_origem_id == arquivo.id,
            Fornecedor.empresa_id == empresa_id,
        ).all():
            saldo_final, saldo_final_tipo, status_pagamento = _calcular_posicao_fornecedor(
                forn.saldo_anterior,
                forn.saldo_anterior_tipo,
                forn.total_debito,
                forn.total_credito,
                bool(forn.lancamentos),
            )
            forn.saldo_final = saldo_final
            forn.saldo_final_tipo = saldo_final_tipo
            forn.valor_a_pagar = saldo_final
            forn.status_pagamento = status_pagamento

        arquivo.status = "CONCLUIDO"
        db.commit()
        logger.info("✅ Processamento concluído para arquivo_id=%d", arquivo_id)

    except Exception:
        logger.exception("Background: erro no arquivo_id=%d empresa_id=%s", arquivo_id, empresa_id)
        if db is not None:
            db.rollback()
            try:
                arquivo = db.query(ArquivoImportado).filter(
                    ArquivoImportado.id == arquivo_id,
                    ArquivoImportado.empresa_id == empresa_id,
                ).first()
                if arquivo:
                    arquivo.status = "ERRO"
                    arquivo.mensagem_erro = (
                        "Não foi possível processar o arquivo. Tente novamente."
                    )
                    db.commit()
            except Exception:
                db.rollback()
                logger.exception(
                    "Falha ao registrar ERRO no arquivo_id=%d empresa_id=%s",
                    arquivo_id,
                    empresa_id,
                )
    finally:
        if db is not None:
            db.close()


# ============================================================================
# UPLOAD E PROCESSAMENTO
# ============================================================================

@router.post("/upload", dependencies=[requer("concilpro.execute"), Depends(require_csrf)])
async def upload_arquivo(
    background_tasks: BackgroundTasks,
    empresa_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Recebe o Razão em PDF ou planilha XLSX/XLS e retorna IMEDIATAMENTE com
    arquivo_id e status=PROCESSANDO. O parsing roda em background; use
    GET /arquivos/{id}/status para polling.

    Prefira XLSX/XLS quando o sistema contábil permitir: a planilha declara o
    que o PDF obriga a inferir (célula tipada, coluna nomeada, sem paginação),
    então o parsing é determinístico e não usa IA.
    """
    from src.domain.concilpro.parser import calcular_hash_arquivo

    try:
        conteudo = await ler_upload_limitado(file)
        hash_arquivo = calcular_hash_arquivo(conteudo)

        # ── Verifica duplicata ──────────────────────────────────────────────────
        result = await db.execute(
            select(ArquivoImportado).where(
                ArquivoImportado.empresa_id == empresa_id,
                ArquivoImportado.hash_arquivo == hash_arquivo,
            )
        )
        existente = result.scalar_one_or_none()

        if existente:
            if existente.status == "PROCESSANDO" and (existente.total_fornecedores or 0) > 0:
                # Processamento em andamento com dados parciais — não interromper
                return {
                    "success": True,
                    "arquivo_id": existente.id,
                    "status": "PROCESSANDO",
                    "message": "Arquivo já está sendo processado.",
                }

            if existente.status == "PROCESSANDO" and (existente.total_fornecedores or 0) == 0:
                # PROCESSANDO com 0 fornecedores = servidor reiniciado durante processamento
                # Trata como ERRO e reprocessa automaticamente
                logger.info(
                    "⚠️ Arquivo id=%d travado em PROCESSANDO (0 fornecedores) — reprocessando…",
                    existente.id,
                )
                # cai no bloco de reprocessamento abaixo

            elif existente.status == "CONCLUIDO" and (existente.total_fornecedores or 0) > 0:
                raise HTTPException(
                    status_code=400,
                    detail="Arquivo já foi importado anteriormente",
                )

            # ERRO ou CONCLUIDO com 0 fornecedores → apaga e reprocessa
            logger.info(
                "♻️ Reprocessando arquivo_id=%d (status=%s, fornecedores=%d)",
                existente.id, existente.status, existente.total_fornecedores or 0,
            )

            # Busca fornecedor IDs para cascade delete
            forn_result = await db.execute(
                select(Fornecedor.id).where(
                    Fornecedor.empresa_id == empresa_id,
                    Fornecedor.arquivo_origem_id == existente.id,
                )
            )
            forn_ids = [r[0] for r in forn_result.all()]

            if forn_ids:
                await db.execute(
                    delete(ConciliacaoInterna).where(
                        ConciliacaoInterna.empresa_id == empresa_id,
                        ConciliacaoInterna.fornecedor_id.in_(forn_ids),
                    )
                )
                await db.execute(
                    delete(Divergencia).where(
                        Divergencia.empresa_id == empresa_id,
                        Divergencia.fornecedor_id.in_(forn_ids),
                    )
                )
                await db.execute(
                    delete(LancamentoFornecedor).where(
                        LancamentoFornecedor.empresa_id == empresa_id,
                        LancamentoFornecedor.fornecedor_id.in_(forn_ids),
                    )
                )
                await db.execute(
                    delete(Fornecedor).where(
                        Fornecedor.empresa_id == empresa_id,
                        Fornecedor.arquivo_origem_id == existente.id,
                    )
                )

            existente.status             = "PROCESSANDO"
            existente.total_fornecedores = 0
            existente.total_lancamentos  = 0
            existente.mensagem_erro      = None
            existente.empresa            = None
            existente.cnpj_empresa       = None
            existente.data_inicio        = None
            existente.data_fim           = None
            # Commit explícito: garante que o registro está visível antes de retornar
            # (get_db faz commit apenas após a resposta ser enviada — race condition)
            await db.commit()

            background_tasks.add_task(
                _processar_arquivo_background,
                existente.id,
                empresa_id,
                conteudo,
            )
            return {
                "success": True,
                "arquivo_id": existente.id,
                "status": "PROCESSANDO",
                "message": "Reprocessamento iniciado — consulte o status para acompanhar.",
            }

        # ── Arquivo novo ────────────────────────────────────────────────────────
        arquivo = ArquivoImportado(
            empresa_id=empresa_id,
            nome_arquivo=file.filename,
            hash_arquivo=hash_arquivo,
            status="PROCESSANDO",
        )
        db.add(arquivo)
        await db.flush()
        await db.refresh(arquivo)
        arquivo_id_novo = arquivo.id
        # Commit explícito: garante que o registro está visível antes de retornar
        # (get_db faz commit apenas após a resposta ser enviada — race condition)
        await db.commit()

        background_tasks.add_task(
            _processar_arquivo_background,
            arquivo_id_novo,
            empresa_id,
            conteudo,
        )

        return {
            "success": True,
            "arquivo_id": arquivo_id_novo,
            "status": "PROCESSANDO",
            "message": "Arquivo recebido. Processamento em andamento — consulte o status para acompanhar.",
        }

    except (HTTPException, AppError):
        # AppError já carrega o status certo (413 no upload acima do limite) —
        # sem esta linha o `except Exception` abaixo o mascararia como 500.
        raise
    except Exception:
        logger.exception("Falha no upload do ConcilPro para empresa_id=%s", empresa_id)
        raise HTTPException(
            status_code=500,
            detail="Não foi possível receber o arquivo para processamento.",
        )


# ============================================================================
# CONSULTAS
# ============================================================================

@router.get("/arquivos/{arquivo_id}/status", dependencies=[requer("concilpro.read")])
async def status_arquivo(
    empresa_id: UUID,
    arquivo_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Retorna o status atual do processamento de um arquivo."""
    result = await db.execute(
        select(ArquivoImportado).where(
            ArquivoImportado.id == arquivo_id,
            ArquivoImportado.empresa_id == empresa_id,
        )
    )
    arquivo = result.scalar_one_or_none()
    if not arquivo:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")
    return {
        "id":                  arquivo.id,
        "status":              arquivo.status,
        "total_fornecedores":  arquivo.total_fornecedores or 0,
        "total_lancamentos":   arquivo.total_lancamentos or 0,
        "mensagem_erro":       arquivo.mensagem_erro,
    }


@router.get("/arquivos", dependencies=[requer("concilpro.read")])
async def listar_arquivos(empresa_id: UUID, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(ArquivoImportado)
        .where(ArquivoImportado.empresa_id == empresa_id)
        .order_by(ArquivoImportado.created_at.desc())
    )
    arquivos = result.scalars().all()
    return [
        {
            "id":                  arq.id,
            "nome_arquivo":        arq.nome_arquivo,
            "status":              arq.status,
            "total_fornecedores":  arq.total_fornecedores,
            "total_lancamentos":   arq.total_lancamentos,
            "periodo_inicio":      arq.data_inicio.isoformat() if arq.data_inicio else None,
            "periodo_fim":         arq.data_fim.isoformat()    if arq.data_fim    else None,
            "created_at":          arq.created_at.isoformat() if arq.created_at else None,
        }
        for arq in arquivos
    ]


@router.get("/resumo/{arquivo_id}", dependencies=[requer("concilpro.read")])
async def obter_resumo(
    empresa_id: UUID,
    arquivo_id: int,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(ArquivoImportado).where(
            ArquivoImportado.id == arquivo_id,
            ArquivoImportado.empresa_id == empresa_id,
        )
    )
    arquivo = result.scalar_one_or_none()
    if not arquivo:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")

    forn_result = await db.execute(
        select(Fornecedor).where(
            Fornecedor.empresa_id == empresa_id,
            Fornecedor.arquivo_origem_id == arquivo_id,
        )
    )
    fornecedores = forn_result.scalars().all()

    return {
        "arquivo": {
            "id":            arquivo.id,
            "nome":          arquivo.nome_arquivo,
            "periodo_inicio": arquivo.data_inicio.isoformat() if arquivo.data_inicio else None,
            "periodo_fim":    arquivo.data_fim.isoformat()    if arquivo.data_fim    else None,
        },
        "estatisticas": {
            "total_fornecedores":        len(fornecedores),
            "total_lancamentos":         arquivo.total_lancamentos,
            "fornecedores_quitados":     sum(1 for f in fornecedores if f.status_pagamento == "QUITADO"),
            "fornecedores_em_aberto":    sum(1 for f in fornecedores if f.status_pagamento == "EM_ABERTO"),
            "fornecedores_adiantados":   sum(1 for f in fornecedores if f.status_pagamento == "ADIANTADO"),
            "fornecedores_sem_movimento": sum(1 for f in fornecedores if f.status_pagamento == "SEM_MOVIMENTO"),
            "fornecedores_com_divergencia": sum(1 for f in fornecedores if f.divergencia_calculo),
            "valor_total_a_pagar":       sum(
                (f.valor_a_pagar or Decimal("0.00") for f in fornecedores),
                Decimal("0.00"),
            ),
        },
    }


@router.get("/fornecedores", dependencies=[requer("concilpro.read")])
async def listar_fornecedores(
    empresa_id: UUID,
    arquivo_id: int,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 500,
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Fornecedor).where(
        Fornecedor.empresa_id == empresa_id,
        Fornecedor.arquivo_origem_id == arquivo_id,
    )

    if status:
        stmt = stmt.where(Fornecedor.status_pagamento == status)

    stmt = stmt.order_by(Fornecedor.valor_a_pagar.desc()).offset(skip).limit(limit)

    result = await db.execute(stmt)
    fornecedores = result.scalars().all()

    return [
        {
            "id":               f.id,
            "codigo_conta":     f.codigo_conta,
            "conta_contabil":   f.conta_contabil,
            "nome_fornecedor":  f.nome_fornecedor,
            "total_credito":    f.total_credito or Decimal("0.00"),
            "total_debito":     f.total_debito or Decimal("0.00"),
            "saldo_final":      f.saldo_final or Decimal("0.00"),
            "valor_a_pagar":    f.valor_a_pagar or Decimal("0.00"),
            "status_pagamento": f.status_pagamento,
            "qtd_nfs_pendentes": f.qtd_nfs_pendentes,
            "qtd_nfs_parciais":  f.qtd_nfs_parciais,
            "divergencia_calculo": f.divergencia_calculo,
        }
        for f in fornecedores
    ]


@router.get("/fornecedores/{fornecedor_id}", dependencies=[requer("concilpro.read")])
async def obter_fornecedor_detalhado(
    empresa_id: UUID,
    fornecedor_id: int,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Fornecedor).where(
            Fornecedor.id == fornecedor_id,
            Fornecedor.empresa_id == empresa_id,
        )
    )
    fornecedor = result.scalar_one_or_none()
    if not fornecedor:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")

    lanc_result = await db.execute(
        select(LancamentoFornecedor)
        .where(
            LancamentoFornecedor.fornecedor_id == fornecedor_id,
            LancamentoFornecedor.empresa_id == empresa_id,
        )
        .order_by(LancamentoFornecedor.data_lancamento)
    )
    lancamentos = lanc_result.scalars().all()

    compras_pendentes = [
        l for l in lancamentos
        if l.tipo_operacao == "COMPRA" and (l.valor_saldo or 0) > 0
    ]

    return {
        "fornecedor": {
            "id":                 fornecedor.id,
            "codigo_conta":       fornecedor.codigo_conta,
            "conta_contabil":     fornecedor.conta_contabil,
            "nome_fornecedor":    fornecedor.nome_fornecedor,
            "cnpj":               fornecedor.cnpj,
            "saldo_anterior":     fornecedor.saldo_anterior or Decimal("0.00"),
            "total_credito":      fornecedor.total_credito or Decimal("0.00"),
            "total_debito":       fornecedor.total_debito or Decimal("0.00"),
            "saldo_final":        fornecedor.saldo_final or Decimal("0.00"),
            "valor_a_pagar":      fornecedor.valor_a_pagar or Decimal("0.00"),
            "status_pagamento":   fornecedor.status_pagamento,
            "divergencia_calculo": fornecedor.divergencia_calculo,
            "divergencia_motivo":  fornecedor.mensagem_erro,
        },
        "compras_pendentes": [
            {
                "id":               c.id,
                "data_lancamento":  c.data_lancamento.isoformat() if c.data_lancamento else None,
                "numero_nf":        c.numero_nf,
                "historico":        c.historico,
                "valor_total":      c.valor_credito or Decimal("0.00"),
                "valor_pago_parcial": c.valor_pago_parcial or Decimal("0.00"),
                "valor_saldo":      c.valor_saldo or Decimal("0.00"),
                "status_pagamento": c.status_pagamento,
            }
            for c in compras_pendentes
        ],
        "todos_lancamentos": [
            {
                "id":            l.id,
                "data":          l.data_lancamento.isoformat() if l.data_lancamento else None,
                "lote":          l.lote,
                "historico":     l.historico,
                "tipo_operacao": l.tipo_operacao,
                "valor_debito":  l.valor_debito or Decimal("0.00"),
                "valor_credito": l.valor_credito or Decimal("0.00"),
                "saldo_apos":    l.saldo_apos_lancamento or Decimal("0.00"),
            }
            for l in lancamentos
        ],
    }


@router.get("/fornecedores/{fornecedor_id}/conciliacao-fifo", dependencies=[requer("concilpro.read")])
async def conciliacao_fifo_detalhada(
    empresa_id: UUID,
    fornecedor_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Retorna compras com trace FIFO de pagamentos."""
    fornecedor_result = await db.execute(
        select(Fornecedor.id).where(
            Fornecedor.id == fornecedor_id,
            Fornecedor.empresa_id == empresa_id,
        )
    )
    if fornecedor_result.scalar_one_or_none() is None:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")

    compras_result = await db.execute(
        select(LancamentoFornecedor)
        .where(
            LancamentoFornecedor.fornecedor_id == fornecedor_id,
            LancamentoFornecedor.empresa_id == empresa_id,
            LancamentoFornecedor.tipo_operacao == "COMPRA",
        )
        .order_by(LancamentoFornecedor.data_lancamento)
    )
    compras = compras_result.scalars().all()

    pags_result = await db.execute(
        select(LancamentoFornecedor)
        .where(
            LancamentoFornecedor.fornecedor_id == fornecedor_id,
            LancamentoFornecedor.empresa_id == empresa_id,
            LancamentoFornecedor.tipo_operacao == "PAGAMENTO",
        )
        .order_by(LancamentoFornecedor.data_lancamento)
    )
    pagamentos_db = pags_result.scalars().all()

    saldo = {c.id: _moeda(c.valor_credito) for c in compras}
    pags_por_nf: dict = {c.id: [] for c in compras}

    for pag in pagamentos_db:
        restante = _moeda(pag.valor_debito)
        for compra in compras:
            if saldo[compra.id] <= 0:
                continue
            if restante <= 0:
                break
            aplicado = min(restante, saldo[compra.id])
            saldo[compra.id] -= aplicado
            restante -= aplicado
            pags_por_nf[compra.id].append({
                "data_pagamento": pag.data_lancamento.isoformat() if pag.data_lancamento else None,
                "historico": pag.historico,
                "valor_pago": aplicado,
                "saldo_restante": saldo[compra.id],
            })

    result = []
    for compra in compras:
        pags = pags_por_nf[compra.id]
        result.append({
            "numero_nf":       compra.numero_nf or "—",
            "data_lancamento": compra.data_lancamento.isoformat() if compra.data_lancamento else None,
            "historico":       compra.historico,
            "valor_total":     compra.valor_credito or Decimal("0.00"),
            "valor_pago":      compra.valor_pago_parcial or Decimal("0.00"),
            "data_pagamento":  pags[-1]["data_pagamento"] if pags else None,
            "valor_saldo":     compra.valor_saldo or Decimal("0.00"),
            "status":          compra.status_pagamento or "PENDENTE",
            "pagamentos":      pags,
        })

    return {"conciliacao": result}


@router.get("/divergencias", dependencies=[requer("concilpro.read")])
async def listar_divergencias(
    empresa_id: UUID,
    arquivo_id: int,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Divergencia)
        .join(Fornecedor, Divergencia.fornecedor_id == Fornecedor.id)
        .where(
            Divergencia.empresa_id == empresa_id,
            Fornecedor.empresa_id == empresa_id,
            Fornecedor.arquivo_origem_id == arquivo_id,
            Divergencia.resolvido.is_(False),
        )
    )
    divergencias = result.scalars().all()
    return [
        {
            "id":            d.id,
            "fornecedor_id": d.fornecedor_id,
            "tipo":          d.tipo,
            "severidade":    d.severidade,
            "descricao":     d.descricao,
            "diferenca":     d.diferenca or Decimal("0.00"),
            "created_at":    d.created_at.isoformat() if d.created_at else None,
        }
        for d in divergencias
    ]


# ============================================================================
# EXPORT
# ============================================================================

@router.get("/export/excel/{arquivo_id}", dependencies=[requer("concilpro.read")])
async def exportar_excel(
    empresa_id: UUID,
    arquivo_id: int,
    tipo: str = Query("completo", pattern="^(completo|em_aberto|divergencias)$"),
    db: AsyncSession = Depends(get_db),
):
    """Exporta dados para Excel (.xlsx) em memória."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    result = await db.execute(
        select(ArquivoImportado).where(
            ArquivoImportado.id == arquivo_id,
            ArquivoImportado.empresa_id == empresa_id,
        )
    )
    arquivo = result.scalar_one_or_none()
    if not arquivo:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliação Fornecedores"

    headers = ["Código", "Conta Contábil", "Fornecedor", "CNPJ",
               "Total Compras", "Total Pagamentos", "Saldo a Pagar",
               "Status", "NFs Pendentes", "Divergência"]

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    stmt = select(Fornecedor).where(
        Fornecedor.empresa_id == empresa_id,
        Fornecedor.arquivo_origem_id == arquivo_id,
    )
    if tipo == "em_aberto":
        stmt = stmt.where(Fornecedor.status_pagamento == "EM_ABERTO")
    elif tipo == "divergencias":
        stmt = stmt.where(Fornecedor.divergencia_calculo.is_(True))

    stmt = stmt.order_by(Fornecedor.valor_a_pagar.desc())
    forn_result = await db.execute(stmt)
    fornecedores_list = forn_result.scalars().all()

    for row, f in enumerate(fornecedores_list, 2):
        ws.cell(row=row, column=1,  value=_celula_texto_segura(f.codigo_conta))
        ws.cell(row=row, column=2,  value=_celula_texto_segura(f.conta_contabil))
        ws.cell(row=row, column=3,  value=_celula_texto_segura(f.nome_fornecedor))
        ws.cell(row=row, column=4,  value=_celula_texto_segura(f.cnpj))
        ws.cell(row=row, column=5,  value=f.total_credito or Decimal("0.00"))
        ws.cell(row=row, column=6,  value=f.total_debito or Decimal("0.00"))
        ws.cell(row=row, column=7,  value=f.valor_a_pagar or Decimal("0.00"))
        ws.cell(row=row, column=8,  value=_celula_texto_segura(f.status_pagamento))
        ws.cell(row=row, column=9,  value=f.qtd_nfs_pendentes)
        ws.cell(row=row, column=10, value="Sim" if f.divergencia_calculo else "Não")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=conciliacao_{tipo}.xlsx"},
    )


@router.get("/export/lancamentos/{arquivo_id}", dependencies=[requer("concilpro.read")])
async def exportar_lancamentos_importacao(
    empresa_id: UUID,
    arquivo_id: int,
    formato: Literal["xlsx", "csv", "txt"] = Query("xlsx"),
    db: AsyncSession = Depends(get_db),
):
    """Exporta os lançamentos do Razão (um por linha original) no layout
    padrão de importação contábil, débito/crédito pareados na mesma linha.

    A conta do fornecedor (`Fornecedor.codigo_conta`) é sempre um dos lados
    do lançamento; o lado é definido pelo próprio `valor_debito`/`valor_credito`
    daquele `CpLancamento`, não pelo `tipo_operacao` — isso cobre COMPRA,
    PAGAMENTO e DEVOLUCAO com a mesma regra, sem precisar de caso especial.
    A contrapartida (`conta_partida`) só existe quando o Razão de origem
    trazia essa coluna; quando ausente, sai em branco — não há cadastro de
    conta bancária/caixa no ConciliaPro para preencher essa lacuna.
    """
    result = await db.execute(
        select(ArquivoImportado).where(
            ArquivoImportado.id == arquivo_id,
            ArquivoImportado.empresa_id == empresa_id,
        )
    )
    if result.scalar_one_or_none() is None:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")

    stmt = (
        select(LancamentoFornecedor, Fornecedor.codigo_conta)
        .join(Fornecedor, LancamentoFornecedor.fornecedor_id == Fornecedor.id)
        .where(
            Fornecedor.empresa_id == empresa_id,
            Fornecedor.arquivo_origem_id == arquivo_id,
        )
        .order_by(LancamentoFornecedor.data_lancamento, Fornecedor.nome_fornecedor)
    )
    lancamentos = (await db.execute(stmt)).all()

    linhas = []
    for lanc, codigo_conta_fornecedor in lancamentos:
        debito_fornecedor = lanc.valor_debito and lanc.valor_debito != 0
        if debito_fornecedor:
            conta_debito, conta_credito = codigo_conta_fornecedor, lanc.conta_partida
            valor = lanc.valor_debito
        else:
            conta_credito, conta_debito = codigo_conta_fornecedor, lanc.conta_partida
            valor = lanc.valor_credito

        linhas.append({
            "Data": lanc.data_lancamento.strftime("%d/%m/%Y"),
            "Cód. Conta Debito": conta_debito or "",
            "Cód. Conta Credito": conta_credito or "",
            "Valor": valor or Decimal("0.00"),
            "Cód. Histórico": "",
            "Complemento Histórico": lanc.historico or "",
            "Inicia Lote": lanc.lote or "",
            "Código Matriz/Filial": "",
            "Centro de Custo Débito": "",
            "Centro de Custo Crédito": "",
        })

    if formato == "csv":
        conteudo = _dicts_to_csv(linhas, COLUNAS_LANCAMENTOS_IMPORTACAO)
        media_type = "text/csv; charset=utf-8"
    elif formato == "txt":
        conteudo = _dicts_to_txt(linhas, COLUNAS_LANCAMENTOS_IMPORTACAO)
        media_type = "text/plain; charset=utf-8"
    else:
        # Estes parâmetros preservam a identidade visual já entregue pelo
        # ConcilPro, embora a estrutura e a escrita sejam compartilhadas.
        conteudo = _dicts_to_xlsx(
            linhas,
            COLUNAS_LANCAMENTOS_IMPORTACAO,
            "Importação Lançamentos",
            # Identidade visual que o escritório já recebe neste relatório;
            # mudá-la seria alteração visível que ninguém pediu.
            cabecalho_cor_fundo="4472C4",
            cabecalho_cor_texto="FFFFFF",
            largura_colunas=22,
        )
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        io.BytesIO(conteudo),
        media_type=media_type,
        headers={
            "Content-Disposition": (
                f"attachment; filename=lancamentos_importacao_{arquivo_id}.{formato}"
            )
        },
    )
