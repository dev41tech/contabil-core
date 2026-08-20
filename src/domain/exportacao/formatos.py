"""Formatação compartilhada dos arquivos de exportação contábil."""

from __future__ import annotations

import csv
import io
from decimal import Decimal

from src.core.errors import ValidationError

# A grafia destes cabeçalhos é contrato com o sistema contábil do escritório.
# Centralizá-los evita que exports de fontes diferentes evoluam fora de sincronia.
COLUNAS_LANCAMENTOS_IMPORTACAO = [
    "Data", "Cód. Conta Debito", "Cód. Conta Credito", "Valor",
    "Cód. Histórico", "Complemento Histórico", "Inicia Lote",
    "Código Matriz/Filial", "Centro de Custo Débito", "Centro de Custo Crédito",
]


def _formatar_valor_br(value: object) -> str:
    """Converte o valor para o padrão decimal exigido pelo TXT do escritório."""
    texto = f"{Decimal(str(value)):f}"
    if "." in texto:
        texto = texto.rstrip("0").rstrip(".")
    return texto.replace(".", ",")


def _celula_segura(value: object) -> object:
    """Neutraliza fórmulas em qualquer texto antes de cruzar CSV/TXT/XLSX."""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _new_workbook(
    title: str,
    colunas: list[str],
    *,
    cabecalho_cor_fundo: str | None = None,
    cabecalho_cor_texto: str | None = None,
    largura_colunas: float | None = None,
):
    """Planilha com cabeçalho na primeira linha.

    O estilo entra como cor em hexadecimal, e não como objeto do openpyxl, para
    que quem chama não precise importar a biblioteca de planilha — era o
    acoplamento que esta extração existe para desfazer. O padrão (só negrito,
    sem preenchimento) reproduz o `ExportacaoService`.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValidationError(message="openpyxl não instalado. Use formato csv.") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(colunas)

    # O padrão reproduz o ExportacaoService. Chamadores com identidade visual
    # própria podem fornecer os estilos sem duplicar a construção da planilha.
    fonte = Font(bold=True, color=cabecalho_cor_texto) if cabecalho_cor_texto else Font(bold=True)
    preenchimento = (
        PatternFill(
            start_color=cabecalho_cor_fundo,
            end_color=cabecalho_cor_fundo,
            fill_type="solid",
        )
        if cabecalho_cor_fundo
        else None
    )
    for cell in ws[1]:
        cell.font = fonte
        if preenchimento is not None:
            cell.fill = preenchimento

    if largura_colunas is not None:
        for indice in range(1, len(colunas) + 1):
            ws.column_dimensions[get_column_letter(indice)].width = largura_colunas

    return wb, ws


def _save_workbook(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _dicts_to_csv(linhas: list[dict], colunas: list[str]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=colunas, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(
        {chave: _celula_segura(valor) for chave, valor in linha.items()}
        for linha in linhas
    )
    return buf.getvalue().encode("utf-8-sig")


def _dicts_to_txt(linhas: list[dict], colunas: list[str]) -> bytes:
    """Serializa o layout do escritório sem cabeçalho e com valor decimal BR.

    Somente ``Valor`` é convertido: códigos de conta podem conter pontos reais
    e não devem ser tratados como números decimais.
    """
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    for linha in linhas:
        row = []
        for coluna in colunas:
            valor = linha.get(coluna, "")
            if coluna == "Valor":
                valor = _formatar_valor_br(valor)
            row.append(_celula_segura(valor))
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def _dicts_to_xlsx(
    linhas: list[dict],
    colunas: list[str],
    sheet_name: str,
    *,
    cabecalho_cor_fundo: str | None = None,
    cabecalho_cor_texto: str | None = None,
    largura_colunas: float | None = None,
) -> bytes:
    wb, ws = _new_workbook(
        sheet_name,
        colunas,
        cabecalho_cor_fundo=cabecalho_cor_fundo,
        cabecalho_cor_texto=cabecalho_cor_texto,
        largura_colunas=largura_colunas,
    )
    for linha in linhas:
        ws.append([_celula_segura(linha.get(coluna, "")) for coluna in colunas])
    return _save_workbook(wb)
