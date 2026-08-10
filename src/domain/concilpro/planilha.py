"""
Parser do Razão de Fornecedores em planilha (XLSX).

Por que existe: no PDF é preciso INFERIR o que a planilha DECLARA. Data vira
texto a parsear, valor vira texto numa posição x, e a coluna (Débito ou
Crédito) só se descobre por aritmética de posição — que foi a origem de uma
família inteira de defeitos: caracteres entrelaçados entre baselines próximas,
valor atribuído à coluna errada, histórico longo quebrado em várias linhas e
lançamento duplicado quando a conta atravessa a quebra de página.

Numa planilha nada disso é possível: a célula tem tipo, a coluna tem nome e o
arquivo não pagina. O parsing é determinístico, sem IA, sem heurística e sem
custo — e o resultado é reproduzível entre execuções.

Medido no razão de 208 fornecedores do CARGO TIME: 208 de 208 blocos fecham
com o "Total da conta" declarado, contra 200 (mais 8 via IA) pelo caminho do
PDF, com 2 lançamentos fabricados e 2 divergências.
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import openpyxl

logger = logging.getLogger(__name__)

# Rótulos aceitos por coluna. O cabeçalho varia entre exports do mesmo sistema
# ("Lote" num razão, "Número" noutro), então a coluna é localizada pelo TEXTO,
# nunca por índice fixo — um índice fixo quebra no primeiro layout diferente.
_ROTULOS: Dict[str, Tuple[str, ...]] = {
    "data":      ("data",),
    "lote":      ("lote", "número", "numero", "nº", "no"),
    "historico": ("histórico", "historico", "descrição", "descricao"),
    "cta":       ("cta.c.part.", "cta c part", "cta.c.part", "contrapartida", "c.part."),
    "debito":    ("débito", "debito"),
    "credito":   ("crédito", "credito"),
    "saldo":     ("saldo-exercício", "saldo-exercicio", "saldo exercício", "saldo"),
}

_TOLERANCIA = Decimal("0.01")


def e_planilha(arquivo_bytes: bytes) -> bool:
    """
    Diz se os bytes são um XLSX.

    XLSX é um ZIP, então não basta o magic 'PK': é preciso conferir que o zip
    contém a estrutura de pasta de trabalho do OpenXML.
    """
    if arquivo_bytes[:2] != b"PK":
        return False
    try:
        import zipfile

        with zipfile.ZipFile(BytesIO(arquivo_bytes)) as z:
            nomes = set(z.namelist())
        return "xl/workbook.xml" in nomes or "[Content_Types].xml" in nomes
    except Exception:
        return False


def _texto(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _decimal(valor) -> Decimal:
    """Converte célula numérica em Decimal com 2 casas."""
    if valor is None or valor == "":
        return Decimal("0")
    if isinstance(valor, str):
        limpo = re.sub(r"[^\d,().+-]", "", valor.strip())
        negativo_parenteses = limpo.startswith("(") and limpo.endswith(")")
        limpo = limpo.strip("()")

        if "," in limpo and "." in limpo:
            # O separador mais à direita é o decimal; o outro é milhar.
            if limpo.rfind(",") > limpo.rfind("."):
                limpo = limpo.replace(".", "").replace(",", ".")
            else:
                limpo = limpo.replace(",", "")
        elif "," in limpo:
            partes = limpo.split(",")
            if len(partes) == 2 and 1 <= len(partes[1]) <= 2:
                limpo = ".".join(partes)
            else:
                limpo = "".join(partes)
        elif "." in limpo:
            partes = limpo.split(".")
            if not (len(partes) == 2 and 1 <= len(partes[1]) <= 2):
                limpo = "".join(partes)

        if negativo_parenteses and not limpo.startswith("-"):
            limpo = "-" + limpo
        try:
            return Decimal(limpo).quantize(Decimal("0.01"))
        except Exception:
            return Decimal("0")
    try:
        return Decimal(str(valor)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0")


def _inteiro_como_texto(valor) -> str:
    """'11190.0' → '11190'; preserva strings como vieram."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    if isinstance(valor, int):
        return str(valor)
    return str(valor).strip()


def localizar_colunas(linhas: List[tuple]) -> Optional[Dict[str, int]]:
    """Acha o índice de cada coluna pelo rótulo do cabeçalho."""
    for linha in linhas[:60]:
        if not linha:
            continue
        rotulos = {
            _texto(v).lower(): i
            for i, v in enumerate(linha)
            if isinstance(v, str) and v.strip()
        }
        if not any(r in rotulos for r in _ROTULOS["debito"]):
            continue
        if not any(r in rotulos for r in _ROTULOS["credito"]):
            continue

        colunas: Dict[str, int] = {}
        for nome, aceitos in _ROTULOS.items():
            for rotulo in aceitos:
                if rotulo in rotulos:
                    colunas[nome] = rotulos[rotulo]
                    break
        return colunas
    return None


def _celula(linha: tuple, colunas: Dict[str, int], nome: str):
    i = colunas.get(nome)
    if i is None or i >= len(linha):
        return None
    return linha[i]


def _extrair_cabecalho(linhas: List[tuple]) -> Dict:
    """Empresa, CNPJ e período, que ficam nas primeiras linhas."""
    info = {"empresa": "", "cnpj": "", "periodo_inicio": None, "periodo_fim": None}
    for linha in linhas[:15]:
        if not linha:
            continue
        celulas = [_texto(v) for v in linha]
        junto = " ".join(c for c in celulas if c)
        if not junto:
            continue
        if junto.startswith("Empresa:") and not info["empresa"]:
            info["empresa"] = next(
                (c for c in celulas[1:] if c and not c.startswith("Folha")), ""
            )
        elif "C.N.P.J." in junto and not info["cnpj"]:
            m = re.search(r"[\d.]{2,}[\d./-]+", junto.split(":", 1)[-1])
            if m:
                info["cnpj"] = m.group().strip()
        elif junto.startswith("Período:"):
            m = re.search(r"(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})", junto)
            if m:
                info["periodo_inicio"] = datetime.strptime(m.group(1), "%d/%m/%Y")
                info["periodo_fim"] = datetime.strptime(m.group(2), "%d/%m/%Y")
    return info


def _nome_do_fornecedor(linha: tuple) -> str:
    """
    Na linha 'Conta:' o nome é a última célula de texto significativa.
    Formato: ['Conta:', 1667, '2.1.3.01.0002', '', '', 'CASSOL MATERIAIS...']
    """
    candidatos = [
        _texto(v) for v in linha[2:]
        if isinstance(v, str) and _texto(v) and not re.fullmatch(r"[\d.]+", _texto(v))
    ]
    return candidatos[-1] if candidatos else ""


def parsear_planilha_razao(arquivo_bytes: bytes) -> Dict:
    """
    Parseia o Razão em XLSX. Mesmo contrato de retorno de parsear_arquivo_razao.
    """
    print("📊 Planilha XLSX detectada — parsing determinístico, sem IA")

    try:
        wb = openpyxl.load_workbook(BytesIO(arquivo_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Não foi possível abrir a planilha: {exc}")

    try:
        ws = wb[wb.sheetnames[0]]
        linhas = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    return _parsear_linhas_razao(arquivo_bytes, linhas)


def parsear_xls_razao(arquivo_bytes: bytes) -> Dict:
    """
    Parseia o Razão no formato legado .xls (Excel 97-2003, binário/BIFF).

    Mesmo contrato de retorno de `parsear_planilha_razao` — só a extração das
    células muda (xlrd em vez de openpyxl), a lógica de negócio é a mesma.
    """
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError(
            "Suporte a .xls (formato antigo) não instalado. Execute: pip install xlrd"
        ) from exc

    print("📊 Planilha XLS (legado) detectada — parsing determinístico, sem IA")

    try:
        wb = xlrd.open_workbook(file_contents=arquivo_bytes)
        ws = wb.sheet_by_index(0)
    except Exception as exc:
        raise ValueError(f"Não foi possível abrir a planilha .xls: {exc}")

    linhas: List[tuple] = [
        tuple(
            _valor_celula_xls(ws.cell(linha_idx, col_idx), wb.datemode)
            for col_idx in range(ws.ncols)
        )
        for linha_idx in range(ws.nrows)
    ]

    return _parsear_linhas_razao(arquivo_bytes, linhas)


def _valor_celula_xls(cell, datemode: int):
    """Converte uma célula xlrd para o mesmo formato que o openpyxl entrega
    (datetime nativo para datas, str para texto, número para o resto)."""
    import xlrd

    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return datetime(*xlrd.xldate_as_tuple(cell.value, datemode))
        except (ValueError, xlrd.xldate.XLDateError):
            return cell.value
    return cell.value


def _parsear_linhas_razao(arquivo_bytes: bytes, linhas: List[tuple]) -> Dict:
    """Lógica de negócio comum às fontes XLSX e XLS: recebe as linhas já
    extraídas (célula tipada, sem paginação) e monta o mesmo contrato de
    retorno de `parsear_arquivo_razao`."""
    # import tardio: parser.py importa este módulo, então importar no topo
    # criaria ciclo
    from src.domain.concilpro.parser import (
        calcular_hash_arquivo, extrair_cnpj, extrair_numero_nf,
    )

    if not linhas:
        raise ValueError("Planilha vazia")

    colunas = localizar_colunas(linhas)
    if not colunas:
        raise ValueError(
            "Cabeçalho não encontrado: a planilha precisa ter colunas 'Débito' e 'Crédito'"
        )
    print(f"   📐 Colunas localizadas pelo cabeçalho: {colunas}")

    info = _extrair_cabecalho(linhas)
    fornecedores: List[Dict] = []
    atual: Optional[Dict] = None
    sem_fechar = 0

    for linha in linhas:
        if not linha:
            continue

        primeiro = _texto(linha[0])
        texto_linha = " ".join(_texto(v) for v in linha if isinstance(v, str))

        # ── Início de um fornecedor ──────────────────────────────────────────
        if primeiro.startswith("Conta:"):
            atual = {
                "codigo_conta": _inteiro_como_texto(linha[1] if len(linha) > 1 else ""),
                "conta_contabil": _texto(linha[2] if len(linha) > 2 else ""),
                "nome_fornecedor": _nome_do_fornecedor(linha),
                "saldo_anterior": Decimal("0"),
                "saldo_anterior_tipo": "",
                "total_debito": None,
                "total_credito": None,
                "lancamentos": [],
                "conciliacao_ia": [],
                "resumo_ia": {},
                "validacao_ia": {},
            }
            fornecedores.append(atual)
            continue

        if atual is None:
            continue

        if "SALDO ANTERIOR" in texto_linha.upper():
            saldo = _decimal(_celula(linha, colunas, "saldo"))
            # A planilha usa sinal (negativo = credor); o resto do sistema
            # trabalha com módulo + tipo, como no PDF.
            atual["saldo_anterior"] = abs(saldo)
            atual["saldo_anterior_tipo"] = "C" if saldo < 0 else ("D" if saldo > 0 else "")
            continue

        if "Total da conta" in texto_linha:
            atual["total_debito"] = _decimal(_celula(linha, colunas, "debito"))
            atual["total_credito"] = _decimal(_celula(linha, colunas, "credito"))
            continue

        # ── Lançamento ───────────────────────────────────────────────────────
        data = _celula(linha, colunas, "data")
        if not isinstance(data, (datetime, date)):
            continue
        if not isinstance(data, datetime):
            data = datetime(data.year, data.month, data.day)

        debito = _decimal(_celula(linha, colunas, "debito"))
        credito = _decimal(_celula(linha, colunas, "credito"))
        if debito == 0 and credito == 0:
            continue

        historico = _texto(_celula(linha, colunas, "historico"))
        saldo = _decimal(_celula(linha, colunas, "saldo"))

        atual["lancamentos"].append({
            "data_lancamento": data,
            "lote": _inteiro_como_texto(_celula(linha, colunas, "lote")),
            "historico": historico,
            "conta_partida": _inteiro_como_texto(_celula(linha, colunas, "cta")) or None,
            "valor_debito": debito,
            "valor_credito": credito,
            "saldo_apos_lancamento": abs(saldo),
            "saldo_tipo": "C" if saldo < 0 else ("D" if saldo > 0 else ""),
            "tipo_operacao": "PAGAMENTO" if debito > 0 else "COMPRA",
            "numero_nf": extrair_numero_nf(historico),
            "cnpj_historico": extrair_cnpj(historico),
            "classificacao_incerta": False,
            "classificado_por_ia": False,
        })

    # ── Fecha os totais e confere contra o declarado ─────────────────────────
    for forn in fornecedores:
        soma_debito = sum(l["valor_debito"] for l in forn["lancamentos"])
        soma_credito = sum(l["valor_credito"] for l in forn["lancamentos"])

        if forn["total_debito"] is None:
            # Sem "Total da conta" na planilha: usa a soma dos lançamentos.
            forn["total_debito"] = soma_debito
            forn["total_credito"] = soma_credito
            continue

        if (abs(soma_debito - forn["total_debito"]) > _TOLERANCIA
                or abs(soma_credito - forn["total_credito"]) > _TOLERANCIA):
            sem_fechar += 1
            logger.warning(
                "⚠️ %s (%s): soma dos lançamentos não fecha com o Total da conta "
                "(D %s vs %s | C %s vs %s)",
                forn["nome_fornecedor"][:40], forn["codigo_conta"],
                soma_debito, forn["total_debito"], soma_credito, forn["total_credito"],
            )

    total_lancamentos = sum(len(f["lancamentos"]) for f in fornecedores)
    print(
        f"✅ Planilha processada: {len(fornecedores)} fornecedores, "
        f"{total_lancamentos} lançamentos, {sem_fechar} não fecham com o total declarado"
    )

    return {
        "hash_arquivo": calcular_hash_arquivo(arquivo_bytes),
        "empresa": info["empresa"],
        "cnpj": info["cnpj"],
        "periodo_inicio": info["periodo_inicio"],
        "periodo_fim": info["periodo_fim"],
        "total_fornecedores": len(fornecedores),
        "total_lancamentos": total_lancamentos,
        "fornecedores": fornecedores,
    }
