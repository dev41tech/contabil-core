"""
Gera a planilha de trabalho com as empresas que precisam de CNPJ real.

Somente leitura — não altera nada em produção. É o passo anterior ao
`conciliar_cnpj.py`: aqui sai a lista para o escritório preencher; lá o CNPJ
volta para o banco.

Contexto: `scripts/import_mrcont.py` nunca leu CNPJ de fonte alguma. A razão
social vem do nome da pasta e o CNPJ é sintetizado do índice
(`_cnpj_placeholder`). Enquanto o CNPJ for placeholder, a exportação fiscal
dessas empresas falha — ela casa o CNPJ da empresa com o emitente da nota.

Execução:
    .venv/Scripts/python listar_cnpj_para_conferencia.py
    .venv/Scripts/python listar_cnpj_para_conferencia.py --saida caminho.xlsx

Colunas da planilha:
    razao_social   — como está no cadastro
    cnpj_atual     — o placeholder que está lá hoje
    cnpj_real      — VAZIO, para o escritório preencher
    origem         — de onde já conseguimos deduzir, quando foi possível
    observacao     — avisos que impedem preenchimento automático
"""
from __future__ import annotations

import argparse
import sys
import unicodedata
from collections import defaultdict

from prod_db import conectar_producao, resumo_destino
from src.core.cnpj import formatar, valido


def normalizar_nome(nome: str) -> str:
    sem_acento = "".join(
        c for c in unicodedata.normalize("NFD", nome or "")
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(sem_acento.upper().split())


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--saida", default="cnpjs_para_conferencia.xlsx")
    args = parser.parse_args()

    print(f"Conectando em {resumo_destino()} (somente leitura)\n")
    conn = conectar_producao()
    cur = conn.cursor()

    cur.execute(
        "SELECT razao_social, cnpj FROM empresas "
        "WHERE deleted_at IS NULL ORDER BY razao_social"
    )
    empresas = cur.fetchall()
    invalidas = [(nome, cnpj) for nome, cnpj in empresas if not valido(cnpj)]

    # Candidatos vindos dos razões já processados pelo ConcilPro
    cur.execute(
        "SELECT DISTINCT empresa, cnpj_empresa FROM cp_arquivo "
        "WHERE cnpj_empresa IS NOT NULL AND empresa IS NOT NULL"
    )
    candidatos: dict[str, set[str]] = defaultdict(set)
    for nome, cnpj in cur.fetchall():
        if valido(cnpj):
            candidatos[normalizar_nome(nome)].add(formatar(cnpj))

    # Nomes repetidos no cadastro impedem casamento automático — existem duas "AXEL"
    ocorrencias: dict[str, int] = defaultdict(int)
    for nome, _ in empresas:
        ocorrencias[normalizar_nome(nome)] += 1

    linhas = []
    for nome, cnpj in invalidas:
        chave = normalizar_nome(nome)
        opcoes = candidatos.get(chave, set())
        sugerido = ""
        origem = ""
        observacao = ""

        if len(opcoes) == 1 and ocorrencias[chave] == 1:
            sugerido = next(iter(opcoes))
            origem = "ConcilPro (razão processado)"
        elif len(opcoes) > 1:
            observacao = f"ConcilPro tem {len(opcoes)} CNPJs para este nome: {sorted(opcoes)}"
        elif ocorrencias[chave] > 1:
            observacao = f"{ocorrencias[chave]} empresas com este mesmo nome no cadastro"
        else:
            observacao = "sem fonte no sistema — buscar em contrato ou certificado A1"

        linhas.append([nome, cnpj, sugerido, origem, observacao])

    print(f"Empresas cadastradas : {len(empresas)}")
    print(f"Com CNPJ inválido    : {len(invalidas)}")
    print(f"Com sugestão do ConcilPro : {sum(1 for l in linhas if l[2])}")
    print()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl não instalado — gerando CSV.")
        import csv
        destino = args.saida.replace(".xlsx", ".csv")
        with open(destino, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            w.writerow(["razao_social", "cnpj_atual", "cnpj_real", "origem", "observacao"])
            w.writerows(linhas)
        print(f"OK: {destino}")
        cur.close()
        conn.close()
        return 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CNPJs a conferir"
    cabecalho = ["Razão Social", "CNPJ atual (placeholder)", "CNPJ real", "Origem da sugestão", "Observação"]
    ws.append(cabecalho)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")

    destaque = PatternFill("solid", fgColor="FFF2CC")
    for linha in linhas:
        ws.append(linha)
        if linha[2]:
            ws.cell(row=ws.max_row, column=3).fill = destaque

    for col, largura in zip("ABCDE", (46, 24, 22, 28, 60)):
        ws.column_dimensions[col].width = largura
    ws.freeze_panes = "A2"

    wb.save(args.saida)
    print(f"OK: {args.saida}")
    print("   Preencha a coluna 'CNPJ real' e devolva — depois rode conciliar_cnpj.py.")

    cur.close()
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
