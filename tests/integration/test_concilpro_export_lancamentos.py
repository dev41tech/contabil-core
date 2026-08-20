"""Testes de integração do export de lançamentos do ConcilPro."""

from __future__ import annotations

import ast
import csv
import inspect
import io
from datetime import date
from decimal import Decimal

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

import src.api.v1.concilpro as modulo_concilpro
import src.domain.exportacao.service as modulo_service
from src.db.models import CpArquivo, CpFornecedor, CpLancamento, Empresa
from src.domain.exportacao.formatos import (
    COLUNAS_LANCAMENTOS_IMPORTACAO,
    _dicts_to_txt,
)
from src.domain.exportacao.service import ExportacaoService


async def _login(client: AsyncClient, tenant, usuario) -> None:
    r = await client.post(
        "/api/v1/auth/login",
        json={"tenant_id": str(tenant.id), "email": usuario.email, "senha": "senha_segura_123"},
    )
    assert r.status_code == 200


async def _setup_arquivo_com_lancamentos(db: AsyncSession, empresa) -> CpArquivo:
    arquivo = CpArquivo(
        empresa_id=empresa.id,
        nome_arquivo="razao.xlsx",
        hash_arquivo="c" * 64,
        status="CONCLUIDO",
    )
    db.add(arquivo)
    await db.flush()

    fornecedor = CpFornecedor(
        empresa_id=empresa.id,
        arquivo_origem_id=arquivo.id,
        codigo_conta="1667",
        conta_contabil="Fornecedores Nacionais",
        nome_fornecedor="ACME LTDA",
    )
    db.add(fornecedor)
    await db.flush()

    db.add_all([
        CpLancamento(
            empresa_id=empresa.id,
            fornecedor_id=fornecedor.id,
            data_lancamento=date(2026, 1, 5),
            lote="99",
            historico="COMPRA NF 123",
            conta_partida="3.1.1",
            valor_debito=Decimal("0.00"),
            valor_credito=Decimal("1000.19"),
            tipo_operacao="COMPRA",
        ),
        CpLancamento(
            empresa_id=empresa.id,
            fornecedor_id=fornecedor.id,
            data_lancamento=date(2026, 1, 10),
            lote=None,
            historico="PAGAMENTO NF 123",
            conta_partida="1.1.B.banco",
            valor_debito=Decimal("1000.19"),
            valor_credito=Decimal("0.00"),
            tipo_operacao="PAGAMENTO",
        ),
    ])
    await db.flush()
    return arquivo


@pytest.mark.asyncio
async def test_exportar_lancamentos_pareia_conta_fornecedor_e_contrapartida(
    client: AsyncClient, db: AsyncSession, tenant, usuario, empresa
):
    await _login(client, tenant, usuario)
    arquivo = await _setup_arquivo_com_lancamentos(db, empresa)

    r = await client.get(
        f"/api/v1/empresas/{empresa.id}/concilpro/export/lancamentos/{arquivo.id}"
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"

    wb = load_workbook_from_bytes(r.content)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == [
        "Data", "Cód. Conta Debito", "Cód. Conta Credito", "Valor",
        "Cód. Histórico", "Complemento Histórico", "Inicia Lote",
        "Código Matriz/Filial", "Centro de Custo Débito", "Centro de Custo Crédito",
    ]

    linhas = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(linhas) == 2

    compra = linhas[0]
    assert compra[0] == "05/01/2026"
    assert compra[1] == "3.1.1"       # débito: contrapartida (conta_partida)
    assert compra[2] == "1667"        # crédito: conta do fornecedor
    # O xlsx guarda número como double — é do formato, não do nosso código.
    # Comparar Decimal exato depois do round-trip só passaria por sorte, com
    # valores representáveis em binário (1000.00 é, 1000.19 não). A precisão
    # decimal exata importa no .txt, e lá o teste é sobre o texto gerado.
    assert Decimal(str(compra[3])) == Decimal("1000.19")
    assert compra[5] == "COMPRA NF 123"
    assert compra[6] == "99"          # Inicia Lote

    pagamento = linhas[1]
    assert pagamento[0] == "10/01/2026"
    assert pagamento[1] == "1667"           # débito: conta do fornecedor
    assert pagamento[2] == "1.1.B.banco"    # crédito: contrapartida
    assert Decimal(str(pagamento[3])) == Decimal("1000.19")
    assert pagamento[6] is None             # sem lote nesse lançamento


@pytest.mark.asyncio
async def test_exportar_lancamentos_arquivo_inexistente_404(
    client: AsyncClient, tenant, usuario, empresa
):
    await _login(client, tenant, usuario)
    r = await client.get(
        f"/api/v1/empresas/{empresa.id}/concilpro/export/lancamentos/999999"
    )
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_exportar_lancamentos_nao_cruza_empresas(
    client: AsyncClient, db: AsyncSession, tenant, usuario, empresa
):
    outra_empresa = Empresa(
        tenant_id=tenant.id,
        razao_social="OUTRA EMPRESA LTDA",
        cnpj="98.765.432/0001-10",
        regime_tributario="simples_nacional",
    )
    db.add(outra_empresa)
    await db.flush()

    arquivo_alheio = await _setup_arquivo_com_lancamentos(db, outra_empresa)

    await _login(client, tenant, usuario)
    r = await client.get(
        f"/api/v1/empresas/{empresa.id}/concilpro/export/lancamentos/{arquivo_alheio.id}"
    )
    assert r.status_code == 404


def load_workbook_from_bytes(content: bytes):
    return load_workbook(io.BytesIO(content), data_only=True)


def test_layout_importacao_tem_uma_unica_fonte_de_colunas():
    """Trava a fonte única porque copiar o layout fez os dois exports divergirem."""
    assert COLUNAS_LANCAMENTOS_IMPORTACAO == [
        "Data", "Cód. Conta Debito", "Cód. Conta Credito", "Valor",
        "Cód. Histórico", "Complemento Histórico", "Inicia Lote",
        "Código Matriz/Filial", "Centro de Custo Débito", "Centro de Custo Crédito",
    ]

    for modulo in (modulo_service, modulo_concilpro):
        arvore = ast.parse(inspect.getsource(modulo))
        nomes_atribuidos = {
            alvo.id
            for no in ast.walk(arvore)
            if isinstance(no, ast.Assign)
            for alvo in no.targets
            if isinstance(alvo, ast.Name)
        }
        assert "COLUNAS_LANCAMENTOS_IMPORTACAO" not in nomes_atribuidos
        assert "_COLUNAS_LANCAMENTOS_IMPORTACAO" not in nomes_atribuidos


@pytest.mark.asyncio
async def test_exportar_lancamentos_txt_compartilha_layout_do_service(
    client: AsyncClient, db: AsyncSession, tenant, usuario, empresa
):
    """Trava a unificação: ConcilPro deve herdar o TXT sem cabeçalho e decimal BR."""
    await _login(client, tenant, usuario)
    arquivo = await _setup_arquivo_com_lancamentos(db, empresa)

    resposta = await client.get(
        f"/api/v1/empresas/{empresa.id}/concilpro/export/lancamentos/{arquivo.id}",
        params={"formato": "txt"},
    )

    linhas_esperadas = [
        {
            "Data": "05/01/2026",
            "Cód. Conta Debito": "3.1.1",
            "Cód. Conta Credito": "1667",
            "Valor": Decimal("1000.19"),
            "Cód. Histórico": "",
            "Complemento Histórico": "COMPRA NF 123",
            "Inicia Lote": "99",
            "Código Matriz/Filial": "",
            "Centro de Custo Débito": "",
            "Centro de Custo Crédito": "",
        },
        {
            "Data": "10/01/2026",
            "Cód. Conta Debito": "1667",
            "Cód. Conta Credito": "1.1.B.banco",
            "Valor": Decimal("1000.19"),
            "Cód. Histórico": "",
            "Complemento Histórico": "PAGAMENTO NF 123",
            "Inicia Lote": "",
            "Código Matriz/Filial": "",
            "Centro de Custo Débito": "",
            "Centro de Custo Crédito": "",
        },
    ]
    # Compara com o serializador compartilhado, não com um método privado do
    # ExportacaoService: é justamente essa fonte única que a refatoração criou,
    # e o teste tem que se apoiar nela para provar isso.
    esperado = _dicts_to_txt(linhas_esperadas, COLUNAS_LANCAMENTOS_IMPORTACAO)

    assert resposta.status_code == 200
    assert resposta.headers["content-type"].startswith("text/plain")
    assert resposta.headers["content-disposition"].endswith(f"_{arquivo.id}.txt")
    assert resposta.content == esperado
    linhas = list(csv.reader(io.StringIO(resposta.content.decode("utf-8-sig")), delimiter=";"))
    assert len(linhas) == 2
    assert linhas[0][0] != "Data"
    assert linhas[0][3] == "1000,19"


@pytest.mark.asyncio
async def test_exportar_lancamentos_xlsx_preserva_aparencia_do_concilpro(
    client: AsyncClient, db: AsyncSession, tenant, usuario, empresa
):
    """Trava azul, texto branco e largura 22 porque são aparência já usada pelo escritório."""
    await _login(client, tenant, usuario)
    arquivo = await _setup_arquivo_com_lancamentos(db, empresa)

    resposta = await client.get(
        f"/api/v1/empresas/{empresa.id}/concilpro/export/lancamentos/{arquivo.id}"
    )
    ws = load_workbook_from_bytes(resposta.content).active

    assert ws["A1"].font.bold is True
    assert ws["A1"].font.color.rgb.endswith("FFFFFF")
    assert ws["A1"].fill.fill_type == "solid"
    assert ws["A1"].fill.fgColor.rgb.endswith("4472C4")
    assert {ws.column_dimensions[letra].width for letra in "ABCDEFGHIJ"} == {22.0}


def test_exportacao_service_xlsx_continua_sem_cabecalho_azul():
    """Trava o padrão simples do serviço para a extração não introduzir estilo visível."""
    conteudo = ExportacaoService(None, None, None)._dicts_to_xlsx(
        [{"Valor": Decimal("1.19")}],
        COLUNAS_LANCAMENTOS_IMPORTACAO,
        "Importação Lançamentos",
    )
    ws = load_workbook_from_bytes(conteudo).active

    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.fill_type is None
