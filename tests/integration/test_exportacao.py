"""Testes de integração — Exportação para ERP."""

from __future__ import annotations

import csv
import io
from datetime import UTC, datetime
from decimal import Decimal
from types import SimpleNamespace
from uuid import uuid4

import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from src.db.models import (
    AgenciaBancaria,
    Comprovante,
    Empresa,
    NotaFiscal,
    PlanoConta,
    Tenant,
    Transacao,
    Usuario,
)
from src.domain.exportacao.service import ExportacaoService
from src.schemas.contabil import ExportJobCreate

_OFX_MINI = """\
OFXHEADER:100
DATA:OFXSGML
VERSION:102
<OFX><BANKMSGSRSV1><STMTTRNRS><STMTRS><BANKTRANLIST>
<STMTTRN>
<TRNTYPE>CREDIT
<DTPOSTED>20240501
<TRNAMT>1200.00
<FITID>EXP_TX1
<MEMO>SERVICO PRESTADO CLIENTE X
</STMTTRN>
</BANKTRANLIST></STMTRS></STMTTRNRS></BANKMSGSRSV1></OFX>
"""


async def _login(client, tenant, usuario) -> str:
    r = await client.post(
        "/api/v1/auth/login",
        json={"tenant_id": str(tenant.id), "email": usuario.email, "senha": "senha_segura_123"},
    )
    assert r.status_code == 200
    return r.json()["csrf_token"]


async def _setup_registros(client, db, empresa, csrf):
    agencia = (
        await client.post(
            f"/api/v1/empresas/{empresa.id}/agencias",
            json={"banco_sigla": "SAFRA", "agencia": "0005", "numero": "55555"},
            headers={"X-CSRF-Token": csrf},
        )
    ).json()

    conta = PlanoConta(empresa_id=empresa.id, codigo="4.1.1", descricao="Serviços Prestados", tipo="receita")
    db.add(conta)
    await db.flush()

    await client.post(
        f"/api/v1/empresas/{empresa.id}/extrato/importar?agencia_id={agencia['id']}",
        files={"arquivo": ("e.ofx", io.BytesIO(_OFX_MINI.encode()), "application/octet-stream")},
        headers={"X-CSRF-Token": csrf},
    )

    await client.post(
        f"/api/v1/empresas/{empresa.id}/regras",
        json={
            "conta_id": str(conta.id),
            "agencia_id": agencia["id"],
            "descricao": "Serviço Prestado",
            "historico": "SERVICO",
            "dc": "C",
            "tipo": "automatica",
        },
        headers={"X-CSRF-Token": csrf},
    )

    await client.post(
        f"/api/v1/empresas/{empresa.id}/neo/processar",
        json={},
        headers={"X-CSRF-Token": csrf},
    )


def _url(empresa_id) -> str:
    return f"/api/v1/empresas/{empresa_id}/exportacao/gerar"


# ── Exportar CSV


@pytest.mark.asyncio
async def test_exportar_csv(client, db, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    await _setup_registros(client, db, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    assert "text/csv" in r.headers["content-type"]
    assert "Content-Disposition" in r.headers
    assert r.headers["Content-Disposition"].endswith(".csv\"")

    # Verifica que tem dados (cabeçalho + pelo menos 1 linha)
    linhas = r.content.decode("utf-8-sig").strip().split("\n")
    assert len(linhas) >= 2  # header + data
    assert "historico" in linhas[0]


@pytest.mark.asyncio
async def test_exportar_xlsx(client, db, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    await _setup_registros(client, db, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "xlsx"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    # Verifica magic bytes do XLSX (PK zip)
    assert r.content[:2] == b"PK"


@pytest.mark.asyncio
async def test_exportar_formato_invalido(client, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    r = await client.post(
        _url(empresa.id),
        json={"formato": "pdf"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 422


@pytest.mark.asyncio
async def test_exportar_sem_csrf_rejeita(client, tenant, usuario, empresa):
    await _login(client, tenant, usuario)
    r = await client.post(_url(empresa.id), json={"formato": "csv"})
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_exportar_sem_registros_retorna_vazio(client, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    # Só cabeçalho, sem linhas de dados
    linhas = r.content.decode("utf-8-sig").strip().split("\n")
    assert len(linhas) == 1  # apenas cabeçalho
    assert r.headers.get("X-Total-Registros") == "0"


# ── Exportar lançamentos (layout de importação)


@pytest.mark.asyncio
async def test_exportar_lancamentos_importacao_pareia_debito_e_credito(
    client, db, tenant, usuario, empresa
):
    csrf = await _login(client, tenant, usuario)
    await _setup_registros(client, db, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv", "tipo": "lancamentos_importacao"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    assert r.headers.get("X-Total-Registros") == "1"

    linhas = list(csv.DictReader(io.StringIO(r.content.decode("utf-8-sig"))))
    assert len(linhas) == 1
    linha = linhas[0]
    assert linha["Data"] == "01/05/2024"
    assert linha["Cód. Conta Credito"] == "4.1.1"
    assert linha["Cód. Conta Debito"] not in ("", None)
    assert Decimal(linha["Valor"]) == Decimal("1200.00")
    assert linha["Complemento Histórico"]
    # Campos sem correspondente no modelo hoje saem em branco.
    assert linha["Cód. Histórico"] == ""
    assert linha["Inicia Lote"] == ""
    assert linha["Código Matriz/Filial"] == ""
    assert linha["Centro de Custo Débito"] == ""
    assert linha["Centro de Custo Crédito"] == ""


@pytest.mark.asyncio
async def test_exportar_lancamentos_importacao_usa_conta_numero_legado(
    db, empresa, usuario
):
    """Este layout importa num sistema contábil externo (legado do MrContador),
    que espera `conta_numero` (ex: 1026) — não o `codigo` hierárquico interno
    do contabil-core (ex: 3.4.2.05.0009). Quando `conta_numero` existe, ele
    tem prioridade sobre `codigo`."""
    conta_debito = PlanoConta(
        empresa_id=empresa.id,
        codigo="3.4.2.05.0009",
        conta_numero=1026,
        descricao="Despesas Bancárias",
        tipo="despesa",
    )
    conta_credito = PlanoConta(
        empresa_id=empresa.id,
        codigo="1.1.02.0001",
        conta_numero=1379,
        descricao="Banco Movimento",
        tipo="ativo",
    )
    db.add_all([conta_debito, conta_credito])
    await db.flush()

    agencia = AgenciaBancaria(empresa_id=empresa.id, banco_sigla="SICREDI", agencia="0001", numero="12345")
    db.add(agencia)
    await db.flush()

    from src.db.models import RegistroContabil

    lancamento_id = uuid4()
    db.add_all([
        RegistroContabil(
            empresa_id=empresa.id, lancamento_id=lancamento_id,
            conta_id=conta_debito.id, agencia_id=agencia.id,
            descricao="Tarifa", historico="TARIFA BANCARIA", historico_extrato="TARIFA BANCARIA",
            dc="D", tipo_regra="manual", valor=Decimal("1.19"),
            data_lancamento=datetime(2026, 1, 5, tzinfo=UTC),
        ),
        RegistroContabil(
            empresa_id=empresa.id, lancamento_id=lancamento_id,
            conta_id=conta_credito.id, agencia_id=agencia.id,
            descricao="Contrapartida bancária: Tarifa", historico="TARIFA BANCARIA",
            historico_extrato="TARIFA BANCARIA",
            dc="C", tipo_regra="manual", valor=Decimal("1.19"),
            data_lancamento=datetime(2026, 1, 5, tzinfo=UTC),
        ),
    ])
    await db.flush()

    service = ExportacaoService(db, empresa.id, usuario.id)
    linhas, _ = await service._exportar_lancamentos_importacao(
        ExportJobCreate(formato="csv"), "csv"
    )
    assert len(linhas) == 1
    assert linhas[0]["Cód. Conta Debito"] == "1026"
    assert linhas[0]["Cód. Conta Credito"] == "1379"


@pytest.mark.asyncio
async def test_exportar_lancamentos_importacao_sem_conta_numero_cai_pro_codigo(
    db, empresa, usuario
):
    """Conta sem `conta_numero` (ex: criada direto no contabil-core, sem ID
    legado do MrContador) cai de volta pro código hierárquico — melhor que
    célula em branco."""
    conta = PlanoConta(
        empresa_id=empresa.id, codigo="3.4.2.05.0009", conta_numero=None,
        descricao="Despesas Bancárias", tipo="despesa",
    )
    contrapartida = PlanoConta(
        empresa_id=empresa.id, codigo="1.1.B.abc123", conta_numero=None,
        descricao="Conta bancária sintética", tipo="ativo",
    )
    db.add_all([conta, contrapartida])
    await db.flush()

    agencia = AgenciaBancaria(empresa_id=empresa.id, banco_sigla="SICREDI", agencia="0002", numero="54321")
    db.add(agencia)
    await db.flush()

    from src.db.models import RegistroContabil

    lancamento_id = uuid4()
    db.add_all([
        RegistroContabil(
            empresa_id=empresa.id, lancamento_id=lancamento_id,
            conta_id=conta.id, agencia_id=agencia.id,
            descricao="Tarifa", historico="TARIFA BANCARIA", historico_extrato="TARIFA BANCARIA",
            dc="D", tipo_regra="manual", valor=Decimal("1.19"),
            data_lancamento=datetime(2026, 1, 5, tzinfo=UTC),
        ),
        RegistroContabil(
            empresa_id=empresa.id, lancamento_id=lancamento_id,
            conta_id=contrapartida.id, agencia_id=agencia.id,
            descricao="Contrapartida bancária: Tarifa", historico="TARIFA BANCARIA",
            historico_extrato="TARIFA BANCARIA",
            dc="C", tipo_regra="manual", valor=Decimal("1.19"),
            data_lancamento=datetime(2026, 1, 5, tzinfo=UTC),
        ),
    ])
    await db.flush()

    service = ExportacaoService(db, empresa.id, usuario.id)
    linhas, _ = await service._exportar_lancamentos_importacao(
        ExportJobCreate(formato="csv"), "csv"
    )
    assert len(linhas) == 1
    assert linhas[0]["Cód. Conta Debito"] == "3.4.2.05.0009"
    assert linhas[0]["Cód. Conta Credito"] == "1.1.B.abc123"


@pytest.mark.asyncio
async def test_exportar_lancamentos_importacao_em_xlsx(client, db, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    await _setup_registros(client, db, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "xlsx", "tipo": "lancamentos_importacao"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"


@pytest.mark.asyncio
async def test_exportar_lancamentos_importacao_ignora_lancamento_sem_par(
    db, empresa, usuario
):
    conta = PlanoConta(empresa_id=empresa.id, codigo="4.1.1", descricao="Serviços", tipo="receita")
    db.add(conta)
    await db.flush()

    agencia = AgenciaBancaria(empresa_id=empresa.id, banco_sigla="ITAU", agencia="0001", numero="12345")
    db.add(agencia)
    await db.flush()

    from src.db.models import RegistroContabil

    db.add(
        RegistroContabil(
            empresa_id=empresa.id,
            lancamento_id=uuid4(),
            conta_id=conta.id,
            agencia_id=agencia.id,
            descricao="Órfão",
            historico="ÓRFÃO SEM PAR",
            historico_extrato="ÓRFÃO SEM PAR",
            dc="D",
            tipo_regra="manual",
            valor=Decimal("50.00"),
            data_lancamento=datetime(2024, 8, 1, tzinfo=UTC),
        )
    )
    await db.flush()

    service = ExportacaoService(db, empresa.id, usuario.id)
    linhas, _ = await service._exportar_lancamentos_importacao(
        ExportJobCreate(formato="csv"), "csv"
    )
    assert linhas == []


# ── Exportar extrato bancário


_OFX_TRES_TX = """\
OFXHEADER:100
DATA:OFXSGML
VERSION:102
<OFX><BANKMSGSRSV1><STMTTRNRS><STMTRS><BANKTRANLIST>
<STMTTRN>
<TRNTYPE>CREDIT
<DTPOSTED>20240501
<TRNAMT>1200.00
<FITID>EXT_TX1
<MEMO>RECEBIMENTO CLIENTE ALFA
</STMTTRN>
<STMTTRN>
<TRNTYPE>DEBIT
<DTPOSTED>20240610
<TRNAMT>-350.00
<FITID>EXT_TX2
<MEMO>PAGAMENTO ENERGIA
</STMTTRN>
<STMTTRN>
<TRNTYPE>DEBIT
<DTPOSTED>20240715
<TRNAMT>-90.00
<FITID>EXT_TX3
<MEMO>TARIFA BANCARIA
</STMTTRN>
</BANKTRANLIST></STMTRS></STMTTRNRS></BANKMSGSRSV1></OFX>
"""


async def _setup_extrato(client, empresa, csrf) -> str:
    """Importa 3 transações (mai, jun, jul de 2024). Retorna o agencia_id."""
    agencia = (
        await client.post(
            f"/api/v1/empresas/{empresa.id}/agencias",
            json={"banco_sigla": "ITAU", "agencia": "7285", "numero": "12287"},
            headers={"X-CSRF-Token": csrf},
        )
    ).json()

    r = await client.post(
        f"/api/v1/empresas/{empresa.id}/extrato/importar?agencia_id={agencia['id']}",
        files={"arquivo": ("e.ofx", io.BytesIO(_OFX_TRES_TX.encode()), "application/octet-stream")},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 201
    return agencia["id"]


def _linhas_csv(resposta) -> list[str]:
    return resposta.content.decode("utf-8-sig").strip().split("\n")


@pytest.mark.asyncio
async def test_conferencia_compara_somas_e_expoe_divergencia_com_direcao(
    db, empresa, usuario
):
    agencia = AgenciaBancaria(
        empresa_id=empresa.id,
        banco_sigla="ITAU",
        agencia="0001",
        numero="12345",
    )
    db.add(agencia)
    await db.flush()

    debito = Transacao(
        empresa_id=empresa.id,
        agencia_id=agencia.id,
        data=datetime(2024, 8, 1, tzinfo=UTC),
        valor=Decimal("10.00"),
        historico="PAGAMENTO DIVERGENTE",
        dc="D",
        hash_dedup="conferencia-debito",
    )
    credito = Transacao(
        empresa_id=empresa.id,
        agencia_id=agencia.id,
        data=datetime(2024, 8, 2, tzinfo=UTC),
        valor=Decimal("125.00"),
        historico="RECEBIMENTO CONCILIADO",
        dc="C",
        hash_dedup="conferencia-credito",
    )
    db.add_all([debito, credito])
    await db.flush()

    db.add_all(
        [
            NotaFiscal(
                empresa_id=empresa.id,
                tipo="nfe",
                numero="DIVERGENTE",
                cnpj_emitente="11.111.111/0001-11",
                valor=Decimal("100000.00"),
                data_emissao=datetime(2024, 8, 1, tzinfo=UTC),
                status="associada",
                transacao_id=debito.id,
                dedup_key="nota-conferencia-debito",
            ),
            Comprovante(
                empresa_id=empresa.id,
                agencia_id=agencia.id,
                transacao_id=debito.id,
                favorecido="FORNECEDOR",
                valor_pago=Decimal("1.00"),
            ),
            NotaFiscal(
                empresa_id=empresa.id,
                tipo="nfse",
                numero="PARTE-1",
                cnpj_emitente=empresa.cnpj,
                valor=Decimal("100.00"),
                data_emissao=datetime(2024, 8, 2, tzinfo=UTC),
                status="associada",
                transacao_id=credito.id,
                dedup_key="nota-conferencia-credito",
            ),
            Comprovante(
                empresa_id=empresa.id,
                agencia_id=agencia.id,
                transacao_id=credito.id,
                favorecido="CLIENTE",
                valor_pago=Decimal("25.00"),
            ),
        ]
    )
    await db.flush()

    service = ExportacaoService(db, empresa.id, usuario.id)
    _, conteudo = await service._exportar_conferencia(ExportJobCreate(formato="csv"), "csv")
    linhas = list(csv.DictReader(io.StringIO(conteudo.decode("utf-8-sig"))))
    principais = {linha["historico_extrato"]: linha for linha in linhas if linha["historico_extrato"]}

    divergente = principais["PAGAMENTO DIVERGENTE"]
    assert divergente["status_conciliacao"] == "Divergente"
    assert Decimal(divergente["total_documentos"]) == Decimal("100001.00")
    assert Decimal(divergente["divergencia_residual"]) == Decimal("-99991.00")

    conciliado = principais["RECEBIMENTO CONCILIADO"]
    assert conciliado["status_conciliacao"] == "Conciliado"
    assert Decimal(conciliado["total_documentos"]) == Decimal("125.00")
    assert Decimal(conciliado["divergencia_residual"]) == Decimal("0.00")


@pytest.mark.parametrize("gatilho", ["=", "+", "-", "@"])
@pytest.mark.asyncio
async def test_exportacao_neutraliza_formula_em_csv_e_xlsx(gatilho, db, empresa, usuario):
    from openpyxl import load_workbook

    service = ExportacaoService(db, empresa.id, usuario.id)
    historico = f"{gatilho}SOMA(A1:A2)"
    linhas = [{"historico": historico}]

    csv_bytes = service._dicts_to_csv(linhas, ["historico"])
    csv_row = next(csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig"))))
    assert csv_row["historico"] == "'" + historico

    xlsx_bytes = service._dicts_to_xlsx(linhas, ["historico"], "Teste")
    workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    assert workbook.active["A2"].value == "'" + historico

    lancamento = SimpleNamespace(
        id=uuid4(),
        data_lancamento=datetime(2026, 8, 1, tzinfo=UTC),
        historico=historico,
        historico_extrato=historico,
        dc="D",
        valor=Decimal("10.00"),
        tipo_regra="manual",
        conta_id=uuid4(),
        agencia_id=uuid4(),
        transacao_id=uuid4(),
    )
    csv_lancamento = next(
        csv.DictReader(
            io.StringIO(service._lancamentos_csv([lancamento]).decode("utf-8-sig"))
        )
    )
    assert csv_lancamento["historico"] == "'" + historico
    assert csv_lancamento["historico_extrato"] == "'" + historico

    workbook = load_workbook(
        io.BytesIO(service._lancamentos_xlsx([lancamento])), data_only=False
    )
    assert workbook.active["C2"].value == "'" + historico
    assert workbook.active["D2"].value == "'" + historico


@pytest.mark.asyncio
async def test_exportar_extrato_traz_as_transacoes(client, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    await _setup_extrato(client, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv", "tipo": "extrato"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    assert r.headers.get("X-Total-Registros") == "3"

    linhas = _linhas_csv(r)
    assert linhas[0].strip() == "data,agencia,historico,dc,valor,status"
    assert len(linhas) == 4  # cabeçalho + 3
    assert "RECEBIMENTO CLIENTE ALFA" in r.text
    assert "ITAU 7285 12287" in r.text, "a agência precisa vir resolvida, não como UUID"


@pytest.mark.asyncio
async def test_exportar_extrato_respeita_o_periodo(client, tenant, usuario, empresa):
    """O arquivo tem que bater com o que a tela mostra — inclusive nos filtros."""
    csrf = await _login(client, tenant, usuario)
    await _setup_extrato(client, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={
            "formato": "csv",
            "tipo": "extrato",
            "data_de": "2024-06-01T00:00:00Z",
            "data_ate": "2024-06-30T23:59:59Z",
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    assert r.headers.get("X-Total-Registros") == "1"
    assert "PAGAMENTO ENERGIA" in r.text
    assert "RECEBIMENTO CLIENTE ALFA" not in r.text


@pytest.mark.asyncio
async def test_exportar_extrato_filtra_por_mes(client, tenant, usuario, empresa):
    """Atalho 'mes' (AAAA-MM) equivale a passar data_de/data_ate do mês inteiro."""
    csrf = await _login(client, tenant, usuario)
    await _setup_extrato(client, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv", "tipo": "extrato", "mes": "2024-06"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    assert r.headers.get("X-Total-Registros") == "1"
    assert "PAGAMENTO ENERGIA" in r.text
    assert "RECEBIMENTO CLIENTE ALFA" not in r.text


@pytest.mark.asyncio
async def test_mes_com_formato_invalido_rejeita(client, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv", "tipo": "extrato", "mes": "junho/2024"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 422


@pytest.mark.asyncio
async def test_exportar_extrato_respeita_o_filtro_de_agencia(client, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    agencia_id = await _setup_extrato(client, empresa, csrf)

    outra = (
        await client.post(
            f"/api/v1/empresas/{empresa.id}/agencias",
            json={"banco_sigla": "BB", "agencia": "0001", "numero": "99999"},
            headers={"X-CSRF-Token": csrf},
        )
    ).json()

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv", "tipo": "extrato", "agencia_id": outra["id"]},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    assert r.headers.get("X-Total-Registros") == "0", "agência sem movimento não traz linha"

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv", "tipo": "extrato", "agencia_id": agencia_id},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.headers.get("X-Total-Registros") == "3"


@pytest.mark.asyncio
async def test_exportar_extrato_respeita_o_filtro_de_status(client, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    await _setup_extrato(client, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv", "tipo": "extrato", "status": "pendente"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.headers.get("X-Total-Registros") == "3"

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv", "tipo": "extrato", "status": "processada"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.headers.get("X-Total-Registros") == "0"


@pytest.mark.asyncio
async def test_exportar_extrato_em_xlsx(client, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    await _setup_extrato(client, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "xlsx", "tipo": "extrato"},
        headers={"X-CSRF-Token": csrf},
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"


@pytest.mark.asyncio
async def test_exportar_extrato_ordena_por_data(client, tenant, usuario, empresa):
    csrf = await _login(client, tenant, usuario)
    await _setup_extrato(client, empresa, csrf)

    r = await client.post(
        _url(empresa.id),
        json={"formato": "csv", "tipo": "extrato"},
        headers={"X-CSRF-Token": csrf},
    )
    datas = [l.split(",")[0] for l in _linhas_csv(r)[1:]]
    assert datas == ["01/05/2024", "10/06/2024", "15/07/2024"]
